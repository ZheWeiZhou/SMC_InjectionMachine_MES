from pymodbus.client import ModbusTcpClient
import struct
import redis
import time
import json
from datetime import datetime
from openpyxl import load_workbook, Workbook
import os
from opcua import Client
from opcua import ua
url="opc.tcp://"+"192.168.1.15:4840"
worker=Client(url)
worker.set_user("localuser1622689641636")
worker.set_password("12345")
worker.connect()
counter = int(worker.get_node("ns=1;i=24").get_value())
current_time = datetime.now().strftime("%Y-%m-%d%H%M%S")
file_name = f'{current_time}.xlsx'
current_curve_a =[]
current_curve_b =[]
current_curve_c =[]
voltage_curve_ab =[]
voltage_curve_bc =[]
voltage_curve_ca =[]
timeindex =[]
powermeter = ModbusTcpClient('192.168.3.5', port=502)
redis_url= '140.135.106.49'
machineID= 'Engel-120'
red= redis.Redis(host=redis_url,port=6379,db=0)
process_activate = False
starttimestemp = ''
endtime = ''

def savedata(item):
    current_curve_a_to_csv = json.dumps(item["current_curve_a"])
    current_curve_b_to_csv = json.dumps(item["current_curve_b"])
    current_curve_c_to_csv = json.dumps(item["current_curve_c"])
    voltage_curve_ab_to_csv = json.dumps(item["voltage_curve_ab"])
    voltage_curve_bc_to_csv = json.dumps(item["voltage_curve_bc"])
    voltage_curve_ca_to_csv = json.dumps(item["voltage_curve_ca"])
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    timeindex_to_csv = json.dumps(item["timeindex"])
    totaltime_to_csv = item["totaltime"]
    motorpower_to_csv = json.dumps(item["motorpower"])
    heaterpower_to_csv = json.dumps(item["heaterpower"])
    plasticmotorenergy_to_csv = item["plasticmotorenergy"]
    closemoldenergy_to_csv = item["closemoldenergy"]
    injection_energy_to_csv = item["injection_energy"]
    row_data = [
        created_at, current_curve_a_to_csv, current_curve_b_to_csv, current_curve_c_to_csv,
        voltage_curve_ab_to_csv,voltage_curve_bc_to_csv,voltage_curve_ca_to_csv,timeindex_to_csv,
        totaltime_to_csv,motorpower_to_csv,heaterpower_to_csv,plasticmotorenergy_to_csv,
        closemoldenergy_to_csv,injection_energy_to_csv
                ]
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Injection_Power"
        # 寫入標題列
        ws.append(["created_at","current_a", "current_b", "current_c", "voltage_ab", "voltage_bc", "voltage_ca",
                   "timeindex","totaltime","motorpower","heaterpower","plasticmotorenergy",
                   "closemoldenergy","injection_energy"
                   ])
    else:
        wb = load_workbook(file_name)
        ws = wb.active
    ws.append(row_data)
    wb.save(file_name)
    print(f"[{created_at}] 數據已成功存入 Excel")  

def decode(high,low):
    raw_bytes = struct.pack('>HH', high, low)
    float_val = struct.unpack('>f', raw_bytes)[0]
    float_val = round(float_val, 2)
    return float_val


while True :
    time.sleep(0.001)
    processstatus= worker.get_node("ns=1;i=77").get_value()
    machinestatus   = red.get(f'{machineID}_status')
    machinestatus   = json.loads(machinestatus)
    currnetcounter = int(worker.get_node("ns=1;i=24").get_value())
    if processstatus == 4:
        if process_activate == False:
            print('DETECT MACHINE WORK')
            starttimestemp = time.perf_counter()
            current_curve_a= []
            current_curve_b= [] 
            current_curve_c =[]
            voltage_curve_ab =[]
            voltage_curve_bc =[]
            voltage_curve_ca =[]
            timeindex=[]
            process_activate = True
        rawbyte = powermeter.read_holding_registers(address= 265, count =30).registers
        currenttime = time.perf_counter()
        timestep = currenttime - starttimestemp
        voltage_ab = decode(rawbyte[0],rawbyte[1])
        voltage_bc = decode(rawbyte[2],rawbyte[3])
        voltage_ca = decode(rawbyte[4],rawbyte[5])
        current_a = decode(rawbyte[24],rawbyte[25])
        current_b = decode(rawbyte[26],rawbyte[27])
        current_c = decode(rawbyte[28],rawbyte[29])
        timeindex.append(timestep)
        current_curve_a.append(current_a)
        current_curve_b.append(current_b)
        current_curve_c.append(current_c)
        voltage_curve_ab.append(voltage_ab)
        voltage_curve_bc.append(voltage_bc)
        voltage_curve_ca.append(voltage_ca)
    if counter<currnetcounter:
        if process_activate == True:
            counter = currnetcounter
            print("DETECT MACHINE FINISH")
            process_activate = False
            endtime = time.perf_counter()
            totaltime = endtime - starttimestemp
            machinepowerinfo   = red.get(f'{machineID}_energy')
            machinepowerinfo   = json.loads(machinepowerinfo)
            motorpower = machinepowerinfo["curve"]["motorpower"]["value"]
            heaterpower = machinepowerinfo["curve"]["heaterpower"]["value"]
            plasticmotorenergy = machinepowerinfo["abstract"]["plasticmotorenergy"]["value"]
            closemoldenergy = machinepowerinfo["abstract"]["closemoldenergy"]["value"]
            injection_energy = machinepowerinfo["abstract"]["injection_energy"]["value"]
            datatocsv = {
                "current_curve_a":current_curve_a,
                "current_curve_b":current_curve_b,
                "current_curve_c":current_curve_c,
                "voltage_curve_ab":voltage_curve_ab,
                "voltage_curve_bc":voltage_curve_bc,
                "voltage_curve_ca":voltage_curve_ca,
                "timeindex":timeindex,
                "totaltime":totaltime,
                "motorpower":motorpower,
                "heaterpower":heaterpower,
                "plasticmotorenergy":plasticmotorenergy,
                "closemoldenergy":closemoldenergy,
                "injection_energy":injection_energy,
            }
            savedata(datatocsv)

